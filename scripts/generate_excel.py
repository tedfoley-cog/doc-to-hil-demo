"""Generate the Excel source documents for the demo."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def create_system_requirements() -> None:
    """Create system_requirements.xlsx with realistic automotive requirements."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "System Requirements"

    headers = ["Req ID", "Description", "ASIL", "Verification", "Subsystem", "Status"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    requirements = [
        ("REQ-SYS-001", "PCM shall manage power mode transitions per state machine spec PCM-SPEC-2024-001", "B", "HIL Test", "PCM", "Approved"),
        ("REQ-SYS-002", "TCM shall execute gear shifts within 400ms total shift time", "B", "HIL Test", "TCM", "Approved"),
        ("REQ-SYS-003", "Emergency stop shall activate within 50ms of critical fault detection", "B", "HIL Test", "PCM", "Approved"),
        ("REQ-SYS-004", "Crank timeout shall prevent starter engagement beyond 10 seconds", "B", "SIL Test", "PCM", "Approved"),
        ("REQ-SYS-005", "Gear range interlock shall prevent reverse above 3 km/h", "B", "HIL Test", "TCM", "Approved"),
        ("REQ-SYS-006", "Battery voltage below 9.0V shall inhibit ACC transition", "B", "SIL Test", "PCM", "Approved"),
        ("REQ-SYS-007", "Torque reduction during shift shall complete within 100ms", "B", "HIL Test", "TCM", "Approved"),
        ("REQ-SYS-008", "DTC P0730 shall be logged on shift abort", "QM", "SIL Test", "TCM", "Approved"),
        ("REQ-SYS-009", "All wheel speed sensors shall report within 2% accuracy", "B", "HIL Test", "ABS", "Approved"),
        ("REQ-SYS-010", "CAN message PCM_Status_1 shall broadcast at 20ms cycle time", "QM", "SIL Test", "PCM", "Approved"),
        ("REQ-SYS-011", "Fuel pump shutoff shall persist in EMERGENCY_STOP until faults cleared", "B", "HIL Test", "PCM", "Approved"),
        ("REQ-SYS-012", "Shift schedule hysteresis band shall be 5-8 km/h to prevent gear hunting", "QM", "SIL Test", "TCM", "Approved"),
        ("REQ-SYS-013", "Park engagement shall be inhibited above 2 km/h", "B", "HIL Test", "TCM", "Approved"),
        ("REQ-SYS-014", "ADAS brake request shall override driver torque request", "D", "HIL Test", "ADAS", "Approved"),
        ("REQ-SYS-015", "Coolant temperature sensor fault shall default to 90C", "QM", "SIL Test", "EMS", "Approved"),
    ]

    for row_idx, req in enumerate(requirements, 2):
        for col_idx, val in enumerate(req, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20
    ws.column_dimensions["B"].width = 70

    out_path = Path("source_documents/excel/system_requirements.xlsx")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Created {out_path}")


def create_test_parameters() -> None:
    """Create test_parameters.xlsx with test conditions and expected values."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Test Parameters"

    headers = ["Test ID", "Requirement", "Parameter", "Condition", "Expected Value", "Tolerance", "Unit"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    params = [
        ("TP-001", "REQ-SYS-003", "Emergency Stop Response", "Critical fault asserted", 50, 5, "ms"),
        ("TP-002", "REQ-SYS-004", "Crank Timeout Duration", "Starter engaged, no RPM", 10000, 100, "ms"),
        ("TP-003", "REQ-SYS-006", "Low Battery Threshold", "Battery voltage decreasing", 9.0, 0.1, "V"),
        ("TP-004", "REQ-SYS-002", "Total Shift Time", "Normal upshift 2->3", 400, 50, "ms"),
        ("TP-005", "REQ-SYS-007", "Torque Reduction Time", "Shift initiated", 100, 10, "ms"),
        ("TP-006", "REQ-SYS-005", "Reverse Inhibit Speed", "N->R requested", 3.0, 0.5, "km/h"),
        ("TP-007", "REQ-SYS-013", "Park Inhibit Speed", "P requested from D", 2.0, 0.3, "km/h"),
        ("TP-008", "REQ-SYS-009", "Wheel Speed Accuracy", "Vehicle at 100 km/h", 2.0, 0.5, "%"),
        ("TP-009", "REQ-SYS-010", "CAN Cycle Time", "Normal operation", 20, 2, "ms"),
        ("TP-010", "REQ-SYS-012", "Shift Hysteresis Band", "Throttle 50%, 3->4 up/down", 7.0, 1.0, "km/h"),
        ("TP-011", "REQ-SYS-014", "AEB Override Torque", "AEB active, driver requesting torque", 0, 5, "Nm"),
        ("TP-012", "REQ-SYS-015", "Coolant Temp Default", "Sensor fault active", 90, 2, "degC"),
    ]

    for row_idx, param in enumerate(params, 2):
        for col_idx, val in enumerate(param, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 22
    ws.column_dimensions["D"].width = 40

    out_path = Path("source_documents/excel/test_parameters.xlsx")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Created {out_path}")


if __name__ == "__main__":
    create_system_requirements()
    create_test_parameters()
    print("Excel files generated successfully")
