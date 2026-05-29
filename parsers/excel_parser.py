"""Parse Excel requirement and test parameter files."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from parsers.models import DocumentType, ParsedDocument, Requirement


def parse_requirements_xlsx(xlsx_path: Path) -> tuple[ParsedDocument, list[Requirement]]:
    """Parse a system requirements Excel workbook."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return (
            ParsedDocument(
                document_id="unknown",
                title="Unknown",
                doc_type=DocumentType.EXCEL_REQ,
                source_path=str(xlsx_path),
            ),
            [],
        )

    doc = ParsedDocument(
        document_id=f"XLSX-{xlsx_path.stem}",
        title=str(ws.title),
        doc_type=DocumentType.EXCEL_REQ,
        source_path=str(xlsx_path),
    )

    requirements: list[Requirement] = []
    header_row = True
    for row in ws.iter_rows(values_only=True):
        if header_row:
            header_row = False
            continue
        if not row or not row[0]:
            continue
        req_id = str(row[0])
        text = str(row[1]) if len(row) > 1 and row[1] else ""
        asil = str(row[2]) if len(row) > 2 and row[2] else "QM"
        verification = str(row[3]) if len(row) > 3 and row[3] else "test"
        requirements.append(
            Requirement(
                req_id=req_id,
                text=text,
                asil=asil,
                verification_method=verification,
                source_document=doc.document_id,
            )
        )

    wb.close()
    return doc, requirements


def parse_test_parameters_xlsx(
    xlsx_path: Path,
) -> tuple[ParsedDocument, list[dict[str, object]]]:
    """Parse a test parameters Excel workbook."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return (
            ParsedDocument(
                document_id="unknown",
                title="Unknown",
                doc_type=DocumentType.EXCEL_PARAMS,
                source_path=str(xlsx_path),
            ),
            [],
        )

    doc = ParsedDocument(
        document_id=f"XLSX-{xlsx_path.stem}",
        title=str(ws.title),
        doc_type=DocumentType.EXCEL_PARAMS,
        source_path=str(xlsx_path),
    )

    headers: list[str] = []
    parameters: list[dict[str, object]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c) if c else f"col_{j}" for j, c in enumerate(row)]
            continue
        if not row or not row[0]:
            continue
        param: dict[str, object] = {}
        for j, val in enumerate(row):
            if j < len(headers):
                param[headers[j]] = val
        parameters.append(param)

    wb.close()
    return doc, parameters
